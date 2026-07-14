import os
import random
import uuid

from openpyxl import load_workbook
from flask import (
    Flask, render_template, request, redirect, url_for, abort, session
)
from sqlalchemy import (
    Column, Integer, MetaData, String, Table, Text, UniqueConstraint,
    create_engine, select,
)

# Templates フォルダ（大文字）を明示的に指定（Windows/他OS どちらでも動くように）
app = Flask(__name__, template_folder="Templates")

# セッション（出題順・得点の保持）の署名に使う鍵。
# 本番では必ず環境変数 SECRET_KEY を設定すること（未設定なら開発用の値を使う）。
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-key-change-in-production")
if not app.debug and app.secret_key == "dev-secret-key-change-in-production":
    app.logger.warning(
        "SECRET_KEY が未設定です。本番では環境変数 SECRET_KEY を必ず設定してください。"
    )

# ---- クイズデータ ----------------------------------------------------------
# 問題と解答は questions.xlsx（Excel）で管理する（一問一答式）。
#
# 1行 = 1問。列（1行目はヘッダー、2行目以降がデータ）：
#   section_id     : URL に使う英数字のID（同じ単元の行には同じIDを書く）
#   section_title  : トップ画面に表示する単元名
#   question       : 問題文
#   answer         : 模範解答（語句など）。表示するだけで自動採点はしない
#   explanation    : 解説（任意。解答表示画面に出る。空欄可）
# 単元の表示順は、行の並び順のとおり。問題の出題順はシャッフルされる。
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
QUESTIONS_FILE = os.path.join(BASE_DIR, "questions.xlsx")

# 出題数の選択肢。単元の問題数を超えるものは表示せず、末尾に「全問」を足す。
COUNT_CHOICES = [10, 30, 50, 100]

# 番号順モードで1回に解く問題数（Excel の上から10問ずつ区切る）
BLOCK_SIZE = 10

# ---- 間違えた問題リスト（苦手リスト） --------------------------------------
# 単元ごとに「あとで解き直したい問題」を貯めておく。
#   - ❌（不正解）を押すと自動で入る
#   - 答え合わせ画面から手動でも追加できる
#   - リスト画面からいつでも削除できる
#   - このリストだけを出題するモードがある
#
# 保存先はデータベース。ブラウザには利用者を見分けるIDだけ持たせる
# （セッションcookie は容量が小さく、問題数が増えると入りきらないため）。
#
# 接続先は環境変数 DATABASE_URL で指定する。
#   本番（Render）   : Neon などの PostgreSQL の接続文字列を環境変数に登録する
#   ローカル（未設定）: 下の SQLite ファイル（review.db）に保存する
# どちらでも同じコードが動くので、開発時に PostgreSQL を用意する必要はない。
#
# 問題は「行番号」ではなく「問題文そのもの」で覚える。こうしておけば Excel に行を
# 挿入・並べ替えしてもリストが壊れない（問題文を書き換えた場合はリストから外れる）。
DATABASE_URL = os.environ.get(
    "DATABASE_URL", "sqlite:///" + os.path.join(BASE_DIR, "review.db")
)
# Render/Heroku 系が渡す postgres:// は SQLAlchemy では postgresql:// と書く必要がある
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

# pool_pre_ping: しばらく放置して切れた接続を掴んだままにしない（無料DBは切断が早い）
engine = create_engine(DATABASE_URL, pool_pre_ping=True, future=True)

metadata = MetaData()
review_items = Table(
    "review_items", metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("user_id", String(64), nullable=False),
    Column("section_id", String(64), nullable=False),
    Column("question", Text, nullable=False),
    # 同じ問題を二重に登録しない
    UniqueConstraint("user_id", "section_id", "question", name="uq_review_item"),
)
metadata.create_all(engine)  # テーブルが無ければ作る（あれば何もしない）


def _user_id():
    """このブラウザを見分けるID（無ければ発行する）"""
    if "uid" not in session:
        session["uid"] = uuid.uuid4().hex
        session.permanent = True  # ブラウザを閉じてもリストが残るように
    return session["uid"]


def get_review(section_id):
    """この単元の間違えた問題リスト（問題文のリスト）を追加順に返す"""
    stmt = (
        select(review_items.c.question)
        .where(review_items.c.user_id == _user_id(),
               review_items.c.section_id == section_id)
        .order_by(review_items.c.id)
    )
    with engine.connect() as conn:
        return [row[0] for row in conn.execute(stmt)]


def add_to_review(section_id, question_text):
    """リストに追加（すでに入っていれば何もしない）"""
    uid = _user_id()
    with engine.begin() as conn:
        exists = conn.execute(
            select(review_items.c.id).where(
                review_items.c.user_id == uid,
                review_items.c.section_id == section_id,
                review_items.c.question == question_text,
            )
        ).first()
        if exists is None:
            conn.execute(review_items.insert().values(
                user_id=uid, section_id=section_id, question=question_text,
            ))


def remove_from_review(section_id, question_text):
    """リストから1問削除する"""
    with engine.begin() as conn:
        conn.execute(review_items.delete().where(
            review_items.c.user_id == _user_id(),
            review_items.c.section_id == section_id,
            review_items.c.question == question_text,
        ))


def clear_review(section_id):
    """この単元のリストを空にする"""
    with engine.begin() as conn:
        conn.execute(review_items.delete().where(
            review_items.c.user_id == _user_id(),
            review_items.c.section_id == section_id,
        ))


def review_questions(section):
    """リストに入っている問題を、いまの Excel の内容と突き合わせて返す。

    返すのは (問題番号（0始まり）, 問題) の組。Excel から消された問題は自然に落ちる。
    """
    saved = get_review(section["id"])
    by_text = {q["question"]: (i, q) for i, q in enumerate(section["questions"])}
    return [by_text[t] for t in saved if t in by_text]


def load_sections():
    """questions.xlsx を読み込んでセクションのリストを返す"""
    wb = load_workbook(QUESTIONS_FILE, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise ValueError("questions.xlsx が空です。")

    # ヘッダー名 → 列番号（列の順番が違っても動くように名前で対応づける）
    index = {}
    for i, name in enumerate(header):
        if name is not None:
            index[str(name).strip()] = i
    required = ["section_id", "section_title", "question", "answer"]
    missing = [c for c in required if c not in index]
    if missing:
        raise ValueError(
            "questions.xlsx のヘッダーに次の列が必要です: " + ", ".join(missing)
        )

    def cell(row, name):
        """存在しない列（explanation 未定義など）にも安全にアクセスする"""
        i = index.get(name)
        return row[i] if i is not None and i < len(row) else None

    sections = []
    by_id = {}
    for row in rows:  # 2行目からがデータ
        # 完全に空の行はスキップ
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue

        sec_id = str(cell(row, "section_id") or "").strip()
        sec_title = str(cell(row, "section_title") or "").strip()
        question = cell(row, "question")
        answer = cell(row, "answer")

        # 必須項目（単元ID・問題文・模範解答）が欠けた行はスキップ
        if not sec_id or question is None or answer is None:
            continue
        question = str(question).strip()
        answer = str(answer).strip()
        if not question or not answer:
            continue

        explanation = cell(row, "explanation")
        explanation = str(explanation).strip() if explanation is not None else ""

        if sec_id not in by_id:
            section = {"id": sec_id, "title": sec_title, "questions": []}
            by_id[sec_id] = section
            sections.append(section)
        by_id[sec_id]["questions"].append({
            "question": question,
            "answer": answer,
            "explanation": explanation,
        })

    wb.close()

    if not sections:
        raise ValueError("questions.xlsx に有効な問題がありません。")
    return sections


# Excel を読み直すためのキャッシュ。
# アクセスのたびに Excel の更新日時(mtime)を確認し、前回から変わっていれば
# 読み直す。変わっていなければキャッシュを使う（膨大な問題でも毎回パースしない）。
_cache = {"mtime": None, "sections": None, "by_id": {}}


def get_sections():
    """最新のセクション一覧を返す（Excel が更新されていれば読み直す）"""
    try:
        mtime = os.path.getmtime(QUESTIONS_FILE)
    except OSError:
        # ファイルが見つからない等。直前の内容があればそれを使う。
        if _cache["sections"] is not None:
            return _cache["sections"]
        raise

    if mtime != _cache["mtime"]:
        try:
            sections = load_sections()
        except Exception as e:
            # Excel を編集中で読めない・書式ミス等。
            # 直前に読めた内容があれば、それで動かし続ける。
            if _cache["sections"] is not None:
                app.logger.warning("questions.xlsx を読み直せませんでした: %s", e)
                return _cache["sections"]
            raise
        _cache["mtime"] = mtime
        _cache["sections"] = sections
        _cache["by_id"] = {s["id"]: s for s in sections}

    return _cache["sections"]


def get_section(section_id):
    """セクションを取得（存在しなければ 404）"""
    get_sections()  # 最新の状態に更新
    section = _cache["by_id"].get(section_id)
    if section is None:
        abort(404)
    return section


def current_quiz(section_id):
    """このセクションで進行中のセッションを返す（無ければ None）"""
    quiz = session.get("quiz")
    if not quiz or quiz.get("section_id") != section_id:
        return None
    return quiz


def current_question(section, quiz_state):
    """いま出題中の問題を返す（全問終了・Excel 変更時は None）"""
    questions = section["questions"]
    pos = quiz_state["pos"]
    if pos >= len(quiz_state["order"]):
        return None
    qidx = quiz_state["order"][pos]
    if qidx >= len(questions):  # Excel が編集されて問題数が減った場合の保険
        return None
    return questions[qidx]


@app.route("/")
def index():
    """トップ画面：セクション（単元）を選ぶ"""
    sections = [
        {"id": s["id"], "title": s["title"], "count": len(s["questions"])}
        for s in get_sections()
    ]
    return render_template("index.html", sections=sections)


def blocks_of(total):
    """Excel の並び順を BLOCK_SIZE 問ずつ区切った一覧を返す（番号は1始まり）"""
    return [
        {"start": s + 1, "end": min(s + BLOCK_SIZE, total)}
        for s in range(0, total, BLOCK_SIZE)
    ]


@app.route("/setup/<section_id>")
def setup(section_id):
    """出題方法を選ぶ画面（ランダム出題の問題数 / 番号順の10問ずつ）"""
    section = get_section(section_id)
    total = len(section["questions"])
    # 用意した選択肢のうち、その単元の問題数を超えないものだけ出す。
    # 最後は必ず「全問」（総数が選択肢と重なる場合は重複させない）。
    counts = [c for c in COUNT_CHOICES if c < total] + [total]
    return render_template(
        "setup.html",
        section_id=section_id,
        section_title=section["title"],
        total=total,
        counts=counts,
        blocks=blocks_of(total),
        review_count=len(review_questions(section)),
    )


@app.route("/block/<section_id>/<int:start_no>")
def block(section_id, start_no):
    """番号順モード：Excel の上から数えて start_no 番から BLOCK_SIZE 問を順番どおり出題"""
    section = get_section(section_id)
    total = len(section["questions"])

    # start_no は1始まりの問題番号。ブロックの先頭（1, 11, 21…）に丸める。
    if not (1 <= start_no <= total):
        return redirect(url_for("setup", section_id=section_id))
    first = (start_no - 1) // BLOCK_SIZE * BLOCK_SIZE  # 0始まりの位置

    order = list(range(first, min(first + BLOCK_SIZE, total)))  # シャッフルしない
    return begin_quiz(section_id, order, block_start=first + 1)


@app.route("/start/<section_id>")
def start(section_id):
    """セクション開始：出題順をシャッフルし、選ばれた問題数だけ出題する"""
    section = get_section(section_id)
    total = len(section["questions"])

    # 出題数（?count=10 など）。未指定・不正な値なら全問。
    try:
        count = int(request.args.get("count", total))
    except ValueError:
        count = total
    count = max(1, min(count, total))  # 1〜総数の範囲に収める

    order = list(range(total))
    random.shuffle(order)  # 出題順をシャッフル
    order = order[:count]  # 先頭 count 問だけを今回の出題にする
    return begin_quiz(section_id, order)


@app.route("/retry/<section_id>")
def retry(section_id):
    """間違えた問題だけをもう一度出題する（結果画面から）"""
    section = get_section(section_id)
    quiz_state = current_quiz(section_id)
    if quiz_state is None:
        return redirect(url_for("setup", section_id=section_id))

    # 直前の回で「不正解」にした問題だけを出題対象にする。
    # Excel が編集されて問題が減った場合に備え、範囲外の番号は捨てる。
    wrong = [i for i in quiz_state.get("wrong", []) if i < len(section["questions"])]
    if not wrong:
        return redirect(url_for("result", section_id=section_id))

    random.shuffle(wrong)
    return begin_quiz(section_id, wrong)


@app.route("/review/<section_id>")
def review(section_id):
    """間違えた問題リストの画面：一覧・削除・このリストから出題"""
    section = get_section(section_id)
    items = [
        {"question": q["question"], "answer": q["answer"]}
        for _, q in review_questions(section)
    ]
    return render_template(
        "review.html",
        section_id=section_id,
        section_title=section["title"],
        items=items,
    )


@app.route("/review/<section_id>/remove", methods=["POST"])
def review_remove(section_id):
    """リストから1問だけ削除する"""
    get_section(section_id)
    remove_from_review(section_id, request.form.get("question", ""))
    return redirect(url_for("review", section_id=section_id))


@app.route("/review/<section_id>/clear", methods=["POST"])
def review_clear(section_id):
    """リストを空にする"""
    get_section(section_id)
    clear_review(section_id)
    return redirect(url_for("review", section_id=section_id))


@app.route("/review/<section_id>/start")
def review_start(section_id):
    """間違えた問題リストだけを出題する"""
    section = get_section(section_id)
    order = [i for i, _ in review_questions(section)]
    if not order:
        return redirect(url_for("review", section_id=section_id))

    random.shuffle(order)
    return begin_quiz(section_id, order)


def begin_quiz(section_id, order, block_start=None):
    """出題リストを受け取り、セッションを初期化してクイズ画面へ送る"""
    session["quiz"] = {
        "section_id": section_id,
        "order": order,
        "pos": 0,      # 現在の出題位置（0始まり）
        "score": 0,    # 自己採点で「正解」にした数
        "wrong": [],   # 「不正解」にした問題の番号（やり直し用）
        # 番号順モードのときだけ、そのブロックの先頭番号（1, 11, 21…）が入る
        "block_start": block_start,
    }
    session.pop("last", None)
    return redirect(url_for("quiz", section_id=section_id))


@app.route("/quiz/<section_id>")
def quiz(section_id):
    """出題画面（問題文＋解答の記入欄）"""
    section = get_section(section_id)
    quiz_state = current_quiz(section_id)
    if quiz_state is None:
        return redirect(url_for("start", section_id=section_id))

    if quiz_state["pos"] >= len(quiz_state["order"]):
        return redirect(url_for("result", section_id=section_id))

    q = current_question(section, quiz_state)
    if q is None:
        return redirect(url_for("start", section_id=section_id))

    return render_template(
        "quiz.html",
        section_id=section_id,
        section_title=section["title"],
        question=q["question"],
        number=quiz_state["pos"] + 1,
        total=len(quiz_state["order"]),
        # 番号順モードでは Excel 上の問題番号（1始まり）も出す
        question_no=(quiz_state["order"][quiz_state["pos"]] + 1
                     if quiz_state.get("block_start") else None),
    )


@app.route("/answer/<section_id>", methods=["POST"])
def answer(section_id):
    """解答を受け取り、答え合わせ画面へリダイレクト（PRG）。ここでは採点しない"""
    section = get_section(section_id)
    quiz_state = current_quiz(section_id)
    if quiz_state is None:
        return redirect(url_for("start", section_id=section_id))

    q = current_question(section, quiz_state)
    if q is None:
        return redirect(url_for("result", section_id=section_id))

    user_answer = (request.form.get("answer") or "").strip()

    # 答え合わせ画面に必要な情報を保存。
    # pos も一緒に持ち、二重採点（戻る／リロード）を防ぐ。
    session["last"] = {
        "section_id": section_id,
        "pos": quiz_state["pos"],
        "question": q["question"],   # 間違えた問題リストへの登録に使う
        "user_answer": user_answer,
        "correct_answer": q["answer"],
        "explanation": q["explanation"],
    }
    session.modified = True

    return redirect(url_for("reveal", section_id=section_id))


@app.route("/reveal/<section_id>")
def reveal(section_id):
    """答え合わせ画面：模範解答と解説を表示し、自分で〇×を押してもらう"""
    section = get_section(section_id)
    quiz_state = current_quiz(section_id)
    last = session.get("last")
    if (quiz_state is None or last is None
            or last.get("section_id") != section_id
            or last.get("pos") != quiz_state["pos"]):
        return redirect(url_for("quiz", section_id=section_id))

    return render_template(
        "reveal.html",
        section_id=section_id,
        section_title=section["title"],
        user_answer=last["user_answer"],
        correct_answer=last["correct_answer"],
        explanation=last["explanation"],
        number=quiz_state["pos"] + 1,
        total=len(quiz_state["order"]),
        # すでに「間違えた問題リスト」に入っているか（手動追加ボタンの表示切替に使う）
        in_review=last["question"] in get_review(section_id),
    )


@app.route("/review/<section_id>/add", methods=["POST"])
def review_add(section_id):
    """答え合わせ画面から、いま出ている問題を手動でリストに追加する"""
    get_section(section_id)
    last = session.get("last")
    if last is None or last.get("section_id") != section_id:
        return redirect(url_for("quiz", section_id=section_id))

    add_to_review(section_id, last["question"])
    return redirect(url_for("reveal", section_id=section_id))


@app.route("/grade/<section_id>", methods=["POST"])
def grade(section_id):
    """自己採点（〇 or ×）を受け取り、次の問題へ進む"""
    get_section(section_id)
    quiz_state = current_quiz(section_id)
    last = session.get("last")
    if (quiz_state is None or last is None
            or last.get("section_id") != section_id
            or last.get("pos") != quiz_state["pos"]):
        # 押し直し・リロードなど。二重に加点せずやり直させる。
        return redirect(url_for("quiz", section_id=section_id))

    if request.form.get("judge") == "correct":
        quiz_state["score"] += 1
    else:
        # 間違えた問題は控えておき、結果画面から復習・やり直しできるようにする
        quiz_state.setdefault("wrong", []).append(quiz_state["order"][quiz_state["pos"]])
        # 「間違えた問題リスト」にも自動で追加する
        add_to_review(section_id, last["question"])

    quiz_state["pos"] += 1  # 次の問題へ進める
    session.pop("last", None)
    session.modified = True

    if quiz_state["pos"] >= len(quiz_state["order"]):
        return redirect(url_for("result", section_id=section_id))
    return redirect(url_for("quiz", section_id=section_id))


@app.route("/result/<section_id>")
def result(section_id):
    """セクション終了：自己採点の結果を表示する"""
    section = get_section(section_id)
    quiz_state = current_quiz(section_id)
    if quiz_state is None:
        return redirect(url_for("start", section_id=section_id))

    score = quiz_state["score"]
    total = len(quiz_state["order"])
    percent = round(score / total * 100) if total else 0

    # 間違えた問題の一覧（復習用）。Excel が編集された場合に備えて範囲外は除く。
    questions = section["questions"]
    wrong = [questions[i] for i in quiz_state.get("wrong", []) if i < len(questions)]

    # 番号順モードなら「この範囲をもう一度」「次の10問へ」を出せるようにする
    block_start = quiz_state.get("block_start")
    block_end = next_block = None
    if block_start:
        block_end = min(block_start + BLOCK_SIZE - 1, len(questions))
        if block_end < len(questions):
            next_block = block_end + 1

    return render_template(
        "result.html",
        section_id=section_id,
        section_title=section["title"],
        score=score,
        total=total,
        percent=percent,
        wrong=wrong,
        block_start=block_start,
        block_end=block_end,
        next_block=next_block,
    )


if __name__ == "__main__":
    # ローカル開発用の簡易サーバ。debug は環境変数で明示的に有効化する。
    #   通常起動      : python app.py
    #   デバッグ起動  : FLASK_DEBUG=1 python app.py  （PowerShell: $env:FLASK_DEBUG=1）
    # 本番では gunicorn 等の WSGI サーバから app を起動する（このブロックは使わない）。
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(debug=debug)
